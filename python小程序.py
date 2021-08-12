from tkinter import *
import requests
from tkinter import messagebox, filedialog
from openpyxl import *

root = Tk()
root.title('信息查询')
root.geometry('550x400')
excel_path_list = ['']
interface_address = Label(root, text='地址(必填)')
interface_address.grid(row=0, column=0, pady=5, sticky=E)
interface_address_loop_text = StringVar()
interface_address_loop = Entry(root, width=50, textvariable=interface_address_loop_text)
interface_address_loop_text.set('https://')
interface_address_loop.grid(row=0, column=1, pady=5, sticky='EW')
id_number = Label(root, text='身份证号码(必填)')
id_number.grid(row=1, column=0, pady=5, sticky=E)
id_number_loop_text = StringVar()
id_number_loop = Entry(root, textvariable=id_number_loop_text)
id_number_loop_text.set(' ')
id_number_loop.grid(row=1, column=1, pady=5, sticky='EW')
full_name = Label(root, text='姓名(必填)')
full_name.grid(row=2, column=0, pady=5, sticky=E)
full_name_loop_text = StringVar()
full_name_loop = Entry(root, textvariable=full_name_loop_text)
full_name_loop_text.set(' ')
full_name_loop.grid(row=2, column=1, pady=5, sticky='EW')
app_key = Label(root, text='秘钥(必填)')
app_key.grid(row=3, column=0, pady=5, sticky=E)
app_key_loop_text = StringVar()
app_key_loop = Entry(root, textvariable=app_key_loop_text)
app_key_loop_text.set('123abc')
app_key_loop.grid(row=3, column=1, pady=5, sticky='EW')


def center_window():
    """窗口居中"""
    screen_height = root.winfo_screenheight()
    screen_width = root.winfo_screenwidth()
    root.update_idletasks()
    root.geometry('%dx%d+%d+%d' % (root.winfo_width(), root.winfo_height(), (screen_width - root.winfo_width()) / 2,
                                   (screen_height - root.winfo_height()) / 2))
    root.deiconify()


center_window()


def select_path():
    """
    选择本地文件
    :return:
    """
    excel_path = filedialog.askopenfilename(title='选择文件', filetypes=[('Excel', '*.xlsx')])
    excel_path_list[0] = excel_path
    root.after(1000, set_file_url())


def get_excel(row, column):
    """
    读取Excel内容
    :param row:
    :param column:
    :return:
    """
    if excel_path_list[0] == '':
        messagebox.askokcancel('提示', '请先选择文件')
    else:
        wb = load_workbook(filename=(excel_path_list[0]))
        names = wb.sheetnames
        sheet = wb[names[0]]
        rows = sheet.max_row
        execl_data = sheet.cell(row=row, column=column).value
        return rows, execl_data


def update_execl(row, value):
    """
    更新Excel
    :return:
    """
    wb = load_workbook(filename=(excel_path_list[0]))
    names = wb.sheetnames
    sheet = wb[names[0]]
    sheet.cell(row=3, column=3, value='是否匹配')
    sheet.cell(row=row, column=3, value=value)
    wb.save(filename=(excel_path_list[0]))


def get_user_info():
    u"""
    单个查询
    :return:
    """
    request_url = interface_address_loop_text.get()
    cardno = id_number_loop_text.get()
    name = full_name_loop_text.get()
    appkey = app_key_loop_text.get()
    if request_url == ' ' or cardno == ' ' or name == ' ' or appkey == ' ':
        messagebox.askokcancel('提示', '参数不能为空')
    else:
        params = {'cardno': cardno,
                  'name': name,
                  'appkey': appkey
                  }
        response = requests.get(request_url, params).json()
        print(response)
        messagebox.askokcancel('查询结果', response)


def get_user_info_all():
    """
    批量查询
    :return:
    """
    response_list = []
    request_url = get_excel(1, 2)[1]
    appkey = get_excel(2, 2)[1]
    for i in range(0, int(get_excel(0, 0)[0]) - 3):
        cardno = get_excel(4 + i, 2)[1]
        name = get_excel(4 + i, 1)[1]
        params = {'cardno': cardno,
                  'name': name,
                  'appkey': appkey}
        if request_url is not None:
            messagebox.askokcancel('提示', '查询中,请稍后')
        response = requests.get(request_url, params).json()
        response_list.append(response)
        update_execl(4 + i, response['result']['resp']['desc'])

    print('查询结果:', response_list)
    messagebox.askokcancel('查询结果', response_list)


def set_file_url():
    file_url = Label(root, text='文件地址(必填)')
    file_url.grid(row=6, column=0, pady=5, sticky=E)
    file_url_loop_text = StringVar()
    file_url_loop = Entry(root, textvariable=file_url_loop_text)
    file_url_loop_text.set(excel_path_list[0])
    file_url_loop.grid(row=6, column=1, pady=5, sticky='EW')


set_file_url()
Button(root, width=20, height=2, text='单个查询', command=get_user_info).grid(row=4, columnspan=2)
Button(root, width=20, height=2, text='选择文件', command=select_path).grid(row=7, columnspan=2)
Button(root, width=20, height=2, text='批量查询', command=get_user_info_all).grid(row=8, columnspan=2)
root.mainloop()
